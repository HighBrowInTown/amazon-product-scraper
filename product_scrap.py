import time
import os
import sys
import argparse
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def get_google_sheets_service():
    """
    Authenticate with Google Sheets API
    Requires: pip install google-auth-oauthlib google-auth-httplib2 google-api-python-client
    """
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
        
        auth.authenticate_user()
        service = build('sheets', 'v4')
        return service
    except ImportError:
        print("❌ Google Sheets libraries not installed!")
        print("   Install with: pip install google-auth-oauthlib google-auth-httplib2 google-api-python-client")
        return None
    except Exception as e:
        print(f"❌ Error authenticating with Google Sheets: {e}")
        return None

def extract_asin_from_url(url):
    """
    Extract ASIN from Amazon URL
    """
    match = re.search(r'/dp/([A-Z0-9]{10})', url)
    if match:
        return match.group(1)
    return None

def scrape_product_details(url):
    """
    Scrape detailed product information from Amazon product page
    Returns dict with rating, bestseller rank, and categories
    """
    options = webdriver.ChromeOptions()
    options.add_argument('--headless=new')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-extensions')
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    driver = None
    product_data = {
        'url': url,
        'asin': extract_asin_from_url(url),
        'title': 'N/A',
        'product_rating': 'N/A',  # Product rating (e.g., 4.1)
        'num_ratings': 'N/A',
        'main_categories': 'N/A',  # CSV format: main_cat1, main_cat2
        'main_categories_rank': 'N/A',  # CSV format: #rank1, #rank2
        'sub_categories': 'N/A',  # CSV format: sub_cat1, sub_cat2
        'sub_categories_rank': 'N/A',  # CSV format: #rank1, #rank2
        'price': 'N/A',
        'status': 'Success'
    }
    
    try:
        print(f"\n🔧 Initializing Chrome WebDriver...")
        driver = webdriver.Chrome(options=options)
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        print(f"🌐 Navigating to: {url}")
        driver.get(url)
        
        # Wait for main content
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.ID, "dp-container")))
        
        time.sleep(2)  # Additional wait for dynamic content
        
        # Extract Title
        try:
            title = driver.find_element(By.ID, "productTitle").text.strip()
            product_data['title'] = title
            print(f"✓ Title: {title[:60]}...")
        except:
            try:
                title = driver.find_element(By.XPATH, "//h1//span[@class='a-size-large product-title-word-break']").text.strip()
                product_data['title'] = title
            except:
                try:
                    title = driver.find_element(By.TAG_NAME, "h1").text.strip()
                    product_data['title'] = title
                except:
                    print("⚠ Could not extract title")
        
        # Extract Rating
        rating_found = False
        try:
            # Method 1: Search page source first (most reliable)
            try:
                page_text = driver.page_source
                rating_match = re.search(r'(\d+\.?\d*)\s+out of 5 stars', page_text)
                if rating_match:
                    rating = rating_match.group(1)
                    product_data['product_rating'] = rating
                    print(f"✓ Product Rating: {rating}")
                    rating_found = True
            except:
                pass
            
            # Method 2: Fallback to element selection if page source didn't work
            if not rating_found:
                rating_selectors = [
                    (By.XPATH, "//a[contains(@class, 'mvt-cm-cr-review-stars-mini-popover')]//span[@class='a-icon-alt']"),
                    (By.XPATH, "//i[contains(@class, 'a-icon-star')]//span[@class='a-icon-alt']"),
                    (By.XPATH, "//span[@aria-label and contains(@aria-label, 'out of')]"),
                ]
                
                for selector_type, selector_value in rating_selectors:
                    try:
                        rating_element = driver.find_element(selector_type, selector_value)
                        rating_text = rating_element.get_attribute('aria-label')
                        if not rating_text:
                            rating_text = rating_element.text
                        
                        if rating_text:
                            rating_match = re.search(r'(\d+\.?\d*)', rating_text)
                            if rating_match:
                                rating = rating_match.group(1)
                                try:
                                    float(rating)
                                    product_data['product_rating'] = rating
                                    print(f"✓ Product Rating: {rating}")
                                    rating_found = True
                                    break
                                except:
                                    continue
                    except:
                        continue
        except Exception as e:
            print(f"Rating extraction error: {e}")
        
        if not rating_found:
            print("⚠ Could not extract product rating")
        
        # Extract Number of Ratings
        num_ratings_found = False
        try:
            # Try multiple selectors for ratings count
            num_ratings_selectors = [
                (By.XPATH, "//span[contains(text(), 'ratings')]"),
                (By.CSS_SELECTOR, "span#acrCustomerReviewText"),
                (By.XPATH, "//span[contains(@class, 'a-size-base') and contains(text(), 'ratings')]"),
                (By.XPATH, "//a[@href='#cm_cr-product_reviews']/span")
            ]
            
            for selector_type, selector_value in num_ratings_selectors:
                try:
                    num_ratings = driver.find_element(selector_type, selector_value).text
                    product_data['num_ratings'] = num_ratings
                    print(f"✓ Number of Ratings: {num_ratings}")
                    num_ratings_found = True
                    break
                except:
                    continue
        except:
            pass
        
        if not num_ratings_found:
            print("⚠ Could not extract number of ratings")
        
        # Extract Price
        try:
            # Try multiple price selectors
            price_selectors = [
                (By.CSS_SELECTOR, "span.a-price-whole"),
                (By.XPATH, "//span[@class='a-price-whole']"),
                (By.XPATH, "//span[contains(@class, 'a-price')]"),
                (By.CSS_SELECTOR, ".a-price.a-text-price.a-size-medium.apexPriceToPay")
            ]
            
            for selector_type, selector_value in price_selectors:
                try:
                    price_elem = driver.find_element(selector_type, selector_value)
                    price = price_elem.text
                    if price:
                        product_data['price'] = price
                        print(f"✓ Price: {price}")
                        break
                except:
                    continue
        except:
            print("⚠ Could not extract price")
        
        # Extract Bestseller Rank and Categories
        try:
            rank_found = False
            
            # Method 1: detailBullets section
            try:
                rank_section = driver.find_element(By.ID, "detailBullets_feature_div")
                rank_text = rank_section.text
                
                # Extract all rank entries (handles multiple categories)
                rank_matches = re.findall(r'#([\d,]+)\s+in\s+([^\n]+)', rank_text)
                
                if rank_matches:
                    # Separate main and sub categories
                    main_cats = []
                    main_ranks = []
                    sub_cats = []
                    sub_ranks = []
                    
                    for i, (rank, cat) in enumerate(rank_matches):
                        cat = cat.strip()
                        # Clean up category name - remove (See Top 100...) or similar text
                        cat = re.sub(r'\s*\(See Top \d+ in[^)]*\)', '', cat).strip()
                        
                        # First entry is main category
                        if i == 0:
                            main_cats.append(cat)
                            main_ranks.append(f"#{rank}")
                            print(f"✓ Main Category: {cat}")
                            print(f"✓ Main Category Rank: #{rank}")
                        else:
                            # Rest are sub-categories
                            sub_cats.append(cat)
                            sub_ranks.append(f"#{rank}")
                    
                    # Store as CSV format
                    if main_cats:
                        product_data['main_categories'] = ", ".join(main_cats)
                        product_data['main_categories_rank'] = ", ".join(main_ranks)
                    
                    if sub_cats:
                        product_data['sub_categories'] = ", ".join(sub_cats)
                        product_data['sub_categories_rank'] = ", ".join(sub_ranks)
                        print(f"✓ Sub Categories: {product_data['sub_categories']}")
                        print(f"✓ Sub Category Ranks: {product_data['sub_categories_rank']}")
                    
                    rank_found = True
            except:
                pass
            
            # Method 2: Look in table rows
            if not rank_found:
                try:
                    table_rows = driver.find_elements(By.XPATH, "//th/span[contains(text(), 'Best Sellers Rank')]")
                    if table_rows:
                        for row in table_rows:
                            try:
                                rank_value = row.find_element(By.XPATH, "./../following-sibling::td").text
                                product_data['main_categories_rank'] = rank_value
                                print(f"✓ Bestseller Rank: {rank_value}")
                                rank_found = True
                                break
                            except:
                                continue
                except:
                    pass
            
            # Method 3: Search in all text
            if not rank_found:
                try:
                    page_text = driver.find_element(By.TAG_NAME, "body").text
                    rank_matches = re.findall(r'#([\d,]+)\s+in\s+([^\n]+)', page_text)
                    
                    if rank_matches:
                        main_cats = []
                        main_ranks = []
                        sub_cats = []
                        sub_ranks = []
                        
                        for i, (rank, cat) in enumerate(rank_matches):
                            cat = cat.strip()
                            # Clean up category name - remove (See Top 100...) or similar text
                            cat = re.sub(r'\s*\(See Top \d+ in[^)]*\)', '', cat).strip()
                            
                            # First entry is main category
                            if i == 0:
                                main_cats.append(cat)
                                main_ranks.append(f"#{rank}")
                                print(f"✓ Main Category (from page text): {cat} (Rank: #{rank})")
                            else:
                                # Rest are sub-categories
                                sub_cats.append(cat)
                                sub_ranks.append(f"#{rank}")
                        
                        # Store as CSV format
                        if main_cats:
                            product_data['main_categories'] = ", ".join(main_cats)
                            product_data['main_categories_rank'] = ", ".join(main_ranks)
                        
                        if sub_cats:
                            product_data['sub_categories'] = ", ".join(sub_cats)
                            product_data['sub_categories_rank'] = ", ".join(sub_ranks)
                            print(f"✓ Sub Categories (from page text): {product_data['sub_categories']}")
                            print(f"✓ Sub Category Ranks (from page text): {product_data['sub_categories_rank']}")
                except:
                    pass
                    
        except:
            print("⚠ Could not extract bestseller rank and categories")
        
        return product_data
    
    except Exception as e:
        print(f"\n❌ Error during scraping: {str(e)}")
        product_data['status'] = f"Error: {str(e)}"
        import traceback
        traceback.print_exc()
        return product_data
    
    finally:
        if driver:
            driver.quit()

def save_to_excel(products_data, output_path):
    """
    Save product data to Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Products"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Add metadata
    ws.merge_cells('A1:J1')
    meta_cell = ws['A1']
    meta_cell.value = "Amazon Product Details"
    meta_cell.font = Font(bold=True, size=14)
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    meta_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    ws.merge_cells('A2:J2')
    date_cell = ws['A2']
    date_cell.value = f"Scraped on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    date_cell.font = Font(italic=True, size=10)
    
    # Add empty row
    ws.append([])
    
    # Add headers
    headers = ["ASIN", "Product Title", "Price", "Product Rating", "# Ratings", "Main Categories", "Main Categories Rank", "Sub Categories", "Sub Categories Rank", "Status"]
    ws.append(headers)
    
    # Style headers
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add product data
    if isinstance(products_data, list):
        products = products_data
    else:
        products = [products_data]
    
    for prod in products:
        ws.append([
            prod.get('asin', 'N/A'),
            prod.get('title', 'N/A'),
            prod.get('price', 'N/A'),
            prod.get('product_rating', 'N/A'),
            prod.get('num_ratings', 'N/A'),
            prod.get('main_categories', 'N/A'),
            prod.get('main_categories_rank', 'N/A'),
            prod.get('sub_categories', 'N/A'),
            prod.get('sub_categories_rank', 'N/A'),
            prod.get('status', 'N/A')
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    
    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 20
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Save file
    wb.save(output_path)
    print(f"✓ Excel file saved to: {output_path}")

def save_to_google_sheets(products_data, spreadsheet_id, sheet_name="Sheet1"):
    """
    Save product data to Google Sheets
    """
    try:
        service = get_google_sheets_service()
        if not service:
            return False
        
        # Prepare data
        if not isinstance(products_data, list):
            products = [products_data]
        else:
            products = products_data
        
        # Create headers
        headers = ["ASIN", "Product Title", "Price", "Product Rating", "# Ratings", "Main Categories", "Main Categories Rank", "Sub Categories", "Sub Categories Rank", "Status"]
        
        # Prepare rows
        rows = [headers]
        for prod in products:
            rows.append([
                prod.get('asin', 'N/A'),
                prod.get('title', 'N/A'),
                prod.get('price', 'N/A'),
                prod.get('product_rating', 'N/A'),
                prod.get('num_ratings', 'N/A'),
                prod.get('main_categories', 'N/A'),
                prod.get('main_categories_rank', 'N/A'),
                prod.get('sub_categories', 'N/A'),
                prod.get('sub_categories_rank', 'N/A'),
                prod.get('status', 'N/A')
            ])
        
        # Write to Google Sheets
        body = {
            'values': rows
        }
        
        result = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{sheet_name}!A1",
            valueInputOption="USER_ENTERED",
            body=body
        ).execute()
        
        print(f"✓ Data saved to Google Sheets: {spreadsheet_id}")
        return True
    except Exception as e:
        print(f"❌ Error saving to Google Sheets: {e}")
        return False

def read_urls_from_file(filepath):
    """
    Read URLs from a file (one URL per line)
    """
    try:
        with open(filepath, 'r') as f:
            urls = [line.strip() for line in f if line.strip() and line.startswith('http')]
        return urls
    except FileNotFoundError:
        print(f"❌ File not found: {filepath}")
        return []
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return []

def main():
    """
    Main function with argument parsing
    """
    parser = argparse.ArgumentParser(
        description='Amazon Product Details Scraper',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python script.py --input "https://www.amazon.in/dp/B0CZL9BM4S"
  python script.py --input "urls.txt"
  python script.py --input "https://www.amazon.in/dp/B0CZL9BM4S" --excel "output.xlsx"
  python script.py --input "urls.txt" --sheets "SPREADSHEET_ID"
        '''
    )
    
    parser.add_argument('--input', required=True, help='URL or file path containing URLs')
    parser.add_argument('--excel', help='Output Excel file path (optional)')
    parser.add_argument('--sheets', help='Google Sheets ID to save data (optional)')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("   🛒 Amazon Product Details Scraper 🛒")
    print("=" * 60)
    
    # Determine if input is URL or file
    urls = []
    if args.input.startswith('http'):
        urls = [args.input]
    else:
        urls = read_urls_from_file(args.input)
    
    if not urls:
        print("❌ No valid URLs found!")
        return
    
    print(f"\n📋 URLs to process: {len(urls)}")
    
    # Scrape products
    all_products = []
    for i, url in enumerate(urls, 1):
        print(f"\n{'='*60}")
        print(f"Processing [{i}/{len(urls)}]")
        print(f"{'='*60}")
        product_data = scrape_product_details(url)
        all_products.append(product_data)
    
    # Save to Excel if specified
    if args.excel:
        save_to_excel(all_products, args.excel)
    else:
        # Default Excel save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_excel = f"amazon_products_{timestamp}.xlsx"
        save_to_excel(all_products, default_excel)
    
    # Save to Google Sheets if specified
    if args.sheets:
        save_to_google_sheets(all_products, args.sheets)
    
    print(f"\n{'='*60}")
    print("✅ SCRAPING COMPLETE!")
    print(f"{'='*60}")
    print(f"📊 Products scraped: {len(all_products)}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠ Scraping interrupted by user.")
    except Exception as e:
        print(f"\n❌ Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()